#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook, Workbook
import random

def random_pick(excel_p, out_p, num):
    '''
    :param excel_p: 原excle表格路径
    :param out_p:  新数据保存的路径
    :param num:  随机生成多少条
    :return:
    '''
    wb = load_workbook(excel_p)
    ws = wb.active

    #原表格的行数
    max_row = ws.max_row
    #原表格的列数
    max_column = ws.max_column

    data_list = random.sample(range(2,max_row+1),num)

    #添加表头行
    data_list.insert(0,1)

    wb_out = Workbook()
    #设置新表格的路径
    wb_out.save(out_p)
    wb_out = load_workbook(out_p)
    ws_out = wb_out.active
    out_row = 1

    for a in data_list:
        for column_a in range(1,max_column+1):
            ws_out.cell(row=out_row, column=column_a).value = ws.cell(row=a, column=column_a).value
        out_row += 1


    wb_out.save(out_p)

    wb.close()
    wb_out.close()

if __name__ == "__main__":
    random_pick(u"E:\\work\\browser/浏览器文章tag相关性标注5000_20171114.xlsx", u"E:\\work\\browser/浏览器文章tag相关性标注5000_20171114_final.xlsx", 2000)